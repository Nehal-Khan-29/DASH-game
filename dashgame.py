import tkinter as tk
from tkinter import *
from tkinter import messagebox
from tkinter import ttk as ttk
from PIL import ImageTk,Image
from datetime import date 
import time
import pandas as pd
import openpyxl
import os
import random
import subprocess
import win32gui
import win32con


# Page Close Confirmations (Messagebox):

def homelogout():
    messagebox.showinfo('Thank You','Logged out successfully')
    home.destroy()
        
def homeclose():
    if messagebox.askokcancel('Quit','Do you want to logout and quit?'):
        home.destroy()
        quit()




    
    
# icon window

icon = tk.Tk()
icon.title('DASH')
icon.iconbitmap("E:\\NK programs\\Python\\python save\\DASH game\\dash img.ico")
image = Image.open("E:\\NK programs\\Python\\python save\\DASH game\\dash intro.png")
tk_image = ImageTk.PhotoImage(image)
image_label = tk.Label(icon, image=tk_image)
image_label.pack()
icon.update()
screen_width = icon.winfo_screenwidth()
screen_height = icon.winfo_screenheight()
window_width = 520  
window_height = 450  
x = int((screen_width - window_width) / 2)
y = int((screen_height - window_height) / 2)
icon.geometry("+{}+{}".format(x, y))
icon.after(2000, icon.destroy)
icon.mainloop()







#Excel add

directory = "E:\\NK programs\\Python\\python save\\DASH game"
filename = "dashpoints.xlsx"
filepathx = os.path.join(directory, filename)

if not os.path.exists(filepathx):
    df = pd.DataFrame({'PLAYER': ["PLAYER 1","PLAYER 2","PLAYER 3","PLAYER 4"," ","WINNER"],'POINTS': [0,0,0,0," "," "]})
    with pd.ExcelWriter(filepathx) as writer:
        df.to_excel(writer, sheet_name='Sheet1', index=False)
    print(f"Excel file '{filename}' created in '{directory}' directory.")
else:
    print(f"Excel file '{filename}' already exists in '{directory}' directory.")
    
    
    
def resetpoints():
    for i in range(2,6):
        wb = openpyxl.load_workbook(filepathx)
        sheet = wb.active
        row_num = i
        col_num = 2
        new_value = 0
        sheet.cell(row=row_num, column=col_num).value = new_value
        sheet.cell(row=7, column=col_num).value = " "
        wb.save(filepathx)

            
            
            
            
            
              
# 1 vs 1:
def start_1v1_game():
    resetpoints()
    process = subprocess.Popen(['python', 'E:\\NK programs\\Python\\python save\\DASH game\\ovo.py'])
    while process.poll() is None:
        time.sleep(0.1)
    pointspage()


    
    
    
    
# 2 vs 2:

def start_2v2_game():
    processs=subprocess.Popen(['python', "E:\\NK programs\\Python\\python save\\DASH game\\tvt.py"])
    while processs.poll() is None:
        time.sleep(0.1)
    pointspage()  
    
    



# 2 vs 2 football:

def start_2v2_footballgame():
    processss=subprocess.Popen(['python', "E:\\NK programs\\Python\\python save\\DASH game\\tfvtf.py"])
    while processss.poll() is None:
        time.sleep(0.1)
    pointspage()  




# instruction page:

def instructionpage():

    instruc=tk.Toplevel()
    instruc.geometry('1366x768')
    instruc.iconbitmap("E:\\NK programs\\Python\\python save\\DASH game\\dash img.ico")
    instruc.title('About')
    instruc.state('zoomed')

    instrucpic=ImageTk.PhotoImage(Image.open("E:\\NK programs\\Python\\python save\\DASH game\\dash img.png"))
    instrucpanel=Label(instruc,image=instrucpic)
    instrucpanel.pack(side='top',fill='both',expand='yes')


    Label(instruc,text=('''
    
    >>> Control keys:
        PLAYER 1 : w , a , s , d
        PLAYER 2 : up , left , down , right
        PLAYER 3 : i , j , k , l
        PLAYER 4 : 8 , 4 , 5 , 6
        
    >>> 1 vs 1
        Find the correct white ball and Try to stay on the white ball to score , Use the basic controls to move
    
    >>> 2 vs 2
        Find the correct white ball and Try to stay on the white ball to score (the total team score decides the winner) , Use the basic controls to move   
    
    >>> 2 vs 2 football
        Try to stay on the white ball to score , Use the basic controls to move
        
        use speacial key for tackle the ball (but reduces the player speed, to once again get the original speed, player need to get(touch) thee ball)
        PLAYER 1 : space
        PLAYER 2 : 0
        PLAYER 3 : P
        PLAYER 4 : 3
        
        '''),font=('Arial',15),bg='white', justify='left').place(x=120,y=170)
    
    instruc.mainloop()
    
    
    
    
    
    
        
    
    
# Points Page:

def pointspage():
    point = tk.Toplevel()
    point.geometry('1366x768')
    point.title('Points')
    point.iconbitmap("E:\\NK programs\\Python\\python save\\DASH game\\dash img.ico")
    point.state('zoomed')

    pointpic = ImageTk.PhotoImage(Image.open("E:\\NK programs\\Python\\python save\\DASH game\\dash img.png"))
    pointpanel = Label(point, image=pointpic)
    pointpanel.pack(side='top', fill='both', expand='yes')

    df = pd.read_excel("E:\\NK programs\\Python\\python save\\DASH game\\dashpoints.xlsx")

    tree = ttk.Treeview(point, column=('#c1', '#c2'), show='headings', height=6)

    tree.column('#1', width=140, minwidth=140, anchor=tk.CENTER)
    tree.column('#2', width=140, minwidth=140, anchor=tk.CENTER)

    tree.heading('#1', text='PLAYER')
    tree.heading('#2', text='POINTS')


    tree.tag_configure("Player1", foreground="red")
    tree.tag_configure("Player2", foreground="green")
    tree.tag_configure("Player3", foreground="blue")
    tree.tag_configure("Player4", foreground="orange")

    tree.pack()

    for index, row in df.iterrows():
        player_tag = "Player{}".format(index + 1)
        tree.insert('', 'end', values=tuple(row), tags=player_tag)

    tree.place(x=580, y=220)

    Button(point, text='CLOSE', font=('Arial', 20), command=point.destroy, height=1, width=18, bg='red',
           fg='white', activebackground='Skyblue', activeforeground='thistle1').place(x=1030, y=580)

    point.mainloop()



            



# About Page:

def aboutpage():

    about=tk.Toplevel()
    about.geometry('1366x768')
    about.iconbitmap("E:\\NK programs\\Python\\python save\\DASH game\\dash img.ico")
    about.title('About')
    about.state('zoomed')

    aboutpic=ImageTk.PhotoImage(Image.open("E:\\NK programs\\Python\\python save\\DASH game\\dash img.png"))
    aboutpanel=Label(about,image=aboutpic)
    aboutpanel.pack(side='top',fill='both',expand='yes')

    def help():
        Label(about,text='''  nehalmicro29@gmail.com  ''',font=('Arial',16)).place(x=750,y=550)

    Button(about,text='Contact Us',font=('Arial',20),command=help,height=1,width=16,bg='white',
    fg='gray6',activebackground='Skyblue',activeforeground='thistle1').place(x=750,y=500)

    Label(about,text=('''This program is created on 17 june 2023 by Nehal Khan'''),font=('Arial',15),bg='white').place(x=250,y=200)

    about.mainloop()








# Home Page:

home=tk.Tk()
home.geometry('1366x768')
home.iconbitmap("E:\\NK programs\\Python\\python save\\DASH game\\dash img.ico")
home.title('DASH')
home.state('zoomed')
home.protocol('WM_DELETE_WINDOW',homeclose)


homepic=ImageTk.PhotoImage(Image.open("E:\\NK programs\\Python\\python save\\DASH game\\dash img.png"))
homepanel=Label(home,image=homepic)
homepanel.pack(side='top',fill='both',expand='yes')


Button(home,text='1 vs 1',font=('Arial',20),command=start_1v1_game,height=1,width=16,bg='blue',
       fg='white',activebackground='Skyblue',activeforeground='thistle1').place(x=460,y=350)
Button(home,text='2 vs 2',font=('Arial',20),command=start_2v2_game,height=1,width=16,bg='blue',
       fg='white',activebackground='Skyblue',activeforeground='thistle1').place(x=820,y=350)
Button(home,text='2 vs 2 football',font=('Arial',20),command=start_2v2_footballgame,height=1,width=16,bg='blue',
       fg='white',activebackground='Skyblue',activeforeground='thistle1').place(x=1180,y=350)


Button(home,text='Player List',font=('Arial',20),command=pointspage,height=1,width=16,bg='blue',
       fg='white',activebackground='Skyblue',activeforeground='thistle1').place(x=100,y=630)
Button(home,text='Instruction',font=('Arial',20,),command=instructionpage,height=1,width=16,bg='blue',
       fg='white',activebackground='Skyblue',activeforeground='thistle1').place(x=460,y=630)
Button(home,text='About Us',font=('Arial',20,),command=aboutpage,height=1,width=16,bg='blue',
       fg='white',activebackground='Skyblue',activeforeground='thistle1').place(x=820,y=630)
Button(home,text='Logout',font=('Arial',20),command=homelogout,height=1,width=16,bg='lightgreen',
       fg='red',activebackground='Skyblue',activeforeground='thistle1').place(x=1180,y=630)



home.mainloop()



